#!/usr/bin/env python3
"""3つの出所から「県 x 業種」の開店数を組み立てる。

「この地域はどの業種が増えるか」を検証するには、業種ごとの開店数が要る。
1つの出所では3業種しか取れず、比べられる組が141しかなくて決着がつかなかった。
出所を足して業種を増やす。

  酒販免許（国税庁・全国47県）        コンビニ / ドラッグ / スーパー
  食品衛生申請等システム（全国157自治体） レストラン / 居酒屋・バー / カフェ / パン屋
  厚生局の医療機関一覧（九州8県）       診療所 / 歯科 / 薬局

期間はそろえる。酒販免許が直近12か月ぶんしか残らないので、
どの出所も**直近12か月に開いたもの**だけを数える。

  .venv-duck/bin/python openings_by_format.py   ->  data/openings_by_format.csv
"""
import csv
import gzip
import io
import re
from collections import Counter
from datetime import date, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
WINDOW_DAYS = 365
CUTOFF = (date.today() - timedelta(days=WINDOW_DAYS)).isoformat()

# 食品衛生の業態 -> 業種列。上から順に見る
FOOD_KIND = [("bar", ("居酒屋", "バー", "スナック", "パブ", "クラブ")),
             ("cafe", ("カフェ", "喫茶", "軽食")),
             ("restaurant", ())]           # 残りは全部レストラン
MOBILE = ("自動車", "露店", "仮設", "臨時", "自動販売機", "行商", "移動")


def clean(s: str) -> str:
    return re.sub(r"^[\s①-⑿\d.、]+", "", (s or "").strip())


def from_liquor() -> Counter:
    """酒販免許。すでに座標化済みのファイルに県コードが入っている。"""
    out = Counter()
    path = DATA / "openings_national_geo.csv"
    m = {"コンビニ": "conv", "ドラッグ・薬局": "drug", "スーパー": "super"}
    for r in csv.DictReader(path.open(encoding="utf-8")):
        col = m.get(r["cat"])
        if col and (r.get("date") or "") >= CUTOFF:
            out[(r["pref_code"], col)] += 1
    return out


def from_food() -> Counter:
    """食品衛生。最新の版から、直近12か月に初回許可が下りたものを数える。"""
    out = Counter()
    snaps = sorted(d for d in (DATA / "food" / "snapshots").glob("*") if d.is_dir())
    src = snaps[-1]
    for path in sorted(list(src.glob("*.csv.gz")) + list(src.glob("*.csv"))):
        raw = path.read_bytes()
        if path.suffix == ".gz":
            raw = gzip.decompress(raw)
        for r in csv.DictReader(io.StringIO(raw.decode("utf-8-sig", errors="replace"))):
            d = (r.get("初回許可年月日") or "").replace("/", "-")
            if d < CUTOFF:
                continue
            pref = (r.get("自治体コード") or "")[1:3]      # 040130 -> 40
            kind = clean(r.get("営業の種類", ""))
            gyotai = clean(r.get("業態", ""))
            if any(w in gyotai for w in MOBILE):
                continue
            if kind.startswith("菓子製造業"):
                out[(pref, "bakery")] += 1
            elif kind.startswith(("飲食店営業", "喫茶店営業")):
                for col, words in FOOD_KIND:
                    if not words or any(w in gyotai for w in words):
                        out[(pref, col)] += 1
                        break
    return out


def wareki_year(s) -> str:
    m = re.match(r"([明大昭平令])\s*(\d+)\.\s*(\d+)\.\s*(\d+)", str(s or "").strip())
    if not m:
        return ""
    base = {"明": 1867, "大": 1911, "昭": 1925, "平": 1988, "令": 2018}[m.group(1)]
    return f"{base + int(m.group(2))}-{int(m.group(3)):02d}-{int(m.group(4)):02d}"


def from_medical() -> Counter:
    """厚生局。指定年月日が開設日にあたる。ファイル名から種別と県を読む。"""
    from prefs import PREFS
    romaji = {r: c for c, r, _ in PREFS} | {"ooita": "44"}   # 大分は ooita 表記
    kind = {"ika": "physician", "shika": "dentist", "yakkyoku": "pharmacy"}
    out = Counter()
    seen = set()
    for path in sorted((DATA / "medical").glob("*.xls*")):
        toks = path.stem.split("_")
        k = next((kind[t] for t in toks if t in kind), None)
        p = next((romaji[t] for t in toks if t in romaji), None)
        if not k or not p or (k, p) in seen:
            continue                       # 同じ県・種別は最新の1版だけ使う
        seen.add((k, p))
        wb = openpyxl.load_workbook(str(path), read_only=True)
        ws = wb[wb.sheetnames[0]]
        for r in ws.iter_rows(min_row=10, values_only=True):
            if r[0] and str(r[0]).strip().isdigit() and len(r) > 7:
                if wareki_year(r[7]) >= CUTOFF:
                    out[(p, k)] += 1
        wb.close()
    return out


def main() -> None:
    print(f"直近12か月＝{CUTOFF} 以降に開いたもの\n")
    parts = {"酒販免許": from_liquor(), "食品衛生": from_food(), "厚生局": from_medical()}
    total = Counter()
    for name, c in parts.items():
        total.update(c)
        prefs = len({p for p, _ in c})
        cols = sorted({f for _, f in c})
        print(f"{name:<10}{sum(c.values()):>7,}件  {prefs:>2}県  {' '.join(cols)}")

    path = DATA / "openings_by_format.csv"
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["pref", "format", "openings"])
        for (p, c), n in sorted(total.items()):
            w.writerow([p, c, n])
    print(f"\n-> {path}  {len(total):,} 行 / 合計 {sum(total.values()):,} 件")


if __name__ == "__main__":
    main()
